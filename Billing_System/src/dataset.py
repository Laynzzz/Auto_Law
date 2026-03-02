"""Master dataset (source-of-truth) schema and operations.

Each law firm has its own dataset at  invoice/{FirmName}/master_{FirmName}.xlsx

v2 format (Phase 20): two sheets — 'cases' and 'appearances'.
  - 'cases': one row per (firm_name, index_number) combination
  - 'appearances': one row per court appearance, linked to cases via case_id

v1 format (legacy): single 'cases' sheet with flat rows.
  - Detected automatically; load_dataset() returns merged dicts either way.

Unique key within a firm file:  (index_number, appearance_date)
"""

import calendar
import contextlib
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import CONFIG_PATH, get_data_root as _cfg_get_data_root, load_config

# ── Schema ────────────────────────────────────────────────────────────

# Original flat columns — kept for backward compatibility.
# All public functions return dicts with these keys (+ new IDs in v2).
COLUMNS = [
    "appearance_date",
    "invoice_number",
    "index_number",
    "case_caption",
    "court",
    "outcome",
    "case_status",
    "charge_amount",
    "invoice_sent_date",
    "paid_status",
    "payment_date",
    "notes",
]

# v2 schema — separate case-level and appearance-level columns
CASE_COLUMNS = [
    "case_id",
    "firm_name",
    "caption",
    "index_number",
    "court",
    "case_status",
    "date_added",
    "notes",
]

APPEARANCE_COLUMNS = [
    "appearance_id",
    "case_id",
    "appearance_date",
    "outcome",
    "outcome_notes",
    "charge_amount",
    "invoice_number",
    "invoice_sent_date",
    "docs_generated_at",
    "draft_created_at",
    "paid_status",
    "payment_date",
    "payment_notes",
]

REQUIRED_COLUMNS = [
    "case_caption",
    "index_number",
    "appearance_date",
    "charge_amount",
]

VALID_CASE_STATUSES = {"Open", "Adjourned", "Closed", "Settled", "Dismissed"}

VALID_PAID_STATUSES = {"Paid", "Unpaid", "Partial"}

# Unique key within a per-firm file
UNIQUE_KEY_COLS = ("index_number", "appearance_date")

# ── Paths ─────────────────────────────────────────────────────────────

PROJECT_ROOT = CONFIG_PATH.parent.parent

_data_root: Path | None = None


def get_data_root() -> Path:
    """Return the data root path (cached after first call)."""
    global _data_root
    if _data_root is None:
        _data_root = _cfg_get_data_root()
    return _data_root


def dataset_path(firm_name: str) -> Path:
    """Return path to a firm's master dataset: invoice/{firm_name}/master_{firm_name}.xlsx"""
    return get_data_root() / "invoice" / firm_name / f"master_{firm_name}.xlsx"


def all_firm_names(config: dict | None = None) -> list[str]:
    """Return list of firm names from config."""
    if config is None:
        config = load_config()
    return [f["name"] for f in config["firms"]]


# ── Format detection ──────────────────────────────────────────────────


def _is_v2_format(wb) -> bool:
    """Return True if the workbook uses the v2 two-sheet format."""
    return "appearances" in wb.sheetnames and "cases" in wb.sheetnames


# ── v2 helpers ────────────────────────────────────────────────────────


def load_cases(firm_name: str) -> list[dict]:
    """Load all rows from the 'cases' sheet (v2 only). Returns list of case dicts."""
    path = dataset_path(firm_name)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {path}")

    wb = load_workbook(path)
    if not _is_v2_format(wb):
        wb.close()
        raise ValueError(f"Dataset for '{firm_name}' is v1 format — migrate first.")

    ws = wb["cases"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def load_appearances(firm_name: str) -> list[dict]:
    """Load all rows from the 'appearances' sheet (v2 only)."""
    path = dataset_path(firm_name)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {path}")

    wb = load_workbook(path)
    if not _is_v2_format(wb):
        wb.close()
        raise ValueError(f"Dataset for '{firm_name}' is v1 format — migrate first.")

    ws = wb["appearances"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def get_case_by_index(firm_name: str, index_number: str, wb=None) -> dict | None:
    """Find a case row by index_number in the v2 cases sheet.

    If wb is provided, uses that workbook (caller manages lifecycle).
    """
    close_wb = False
    if wb is None:
        path = dataset_path(firm_name)
        wb = load_workbook(path)
        close_wb = True

    try:
        if not _is_v2_format(wb):
            return None
        ws = wb["cases"]
        headers = [cell.value for cell in ws[1]]
        target = index_number.strip().lower()
        idx_col = headers.index("index_number")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            if str(row[idx_col] or "").strip().lower() == target:
                return dict(zip(headers, row))
        return None
    finally:
        if close_wb:
            wb.close()


def get_or_create_case_id(
    firm_name: str, index_number: str, wb=None
) -> str:
    """Return the case_id for the given index_number, creating a new case row if needed.

    If wb is provided, the caller must save the workbook after this call.
    """
    existing = get_case_by_index(firm_name, index_number, wb=wb)
    if existing:
        return existing["case_id"]
    return str(uuid.uuid4())


def _merge_case_appearance(case_dict: dict, app_dict: dict) -> dict:
    """Merge a case row and an appearance row into a backward-compatible flat dict.

    Mapping:
      cases.caption       -> merged as case_caption
      cases.notes         -> merged as notes
      appearances.outcome_notes -> merged as outcome_notes (new field)
      appearances.payment_notes -> merged as payment_notes (new field)
    """
    merged = {}

    # Appearance-level fields first (these are the "row" in the flat model)
    merged["appearance_id"] = app_dict.get("appearance_id")
    merged["case_id"] = app_dict.get("case_id")
    merged["appearance_date"] = app_dict.get("appearance_date")
    merged["invoice_number"] = app_dict.get("invoice_number")
    merged["outcome"] = app_dict.get("outcome")
    merged["outcome_notes"] = app_dict.get("outcome_notes")
    merged["charge_amount"] = app_dict.get("charge_amount")
    merged["invoice_sent_date"] = app_dict.get("invoice_sent_date")
    merged["docs_generated_at"] = app_dict.get("docs_generated_at")
    merged["draft_created_at"] = app_dict.get("draft_created_at")
    merged["paid_status"] = app_dict.get("paid_status")
    merged["payment_date"] = app_dict.get("payment_date")
    merged["payment_notes"] = app_dict.get("payment_notes")

    # Case-level fields
    merged["index_number"] = case_dict.get("index_number")
    merged["case_caption"] = case_dict.get("caption")
    merged["court"] = case_dict.get("court")
    merged["case_status"] = case_dict.get("case_status")
    merged["firm_name"] = case_dict.get("firm_name")
    merged["date_added"] = case_dict.get("date_added")
    merged["notes"] = case_dict.get("notes")

    return merged


def _split_row_data(row_data: dict, firm_name: str) -> tuple[dict, dict]:
    """Split a flat row_data dict into (case_fields, appearance_fields).

    Maps backward-compatible keys to the v2 schema:
      case_caption -> caption
      notes        -> notes (case-level)
    """
    case_fields: dict = {}
    app_fields: dict = {}

    # Map flat keys to case fields
    _case_key_map = {
        "case_caption": "caption",
        "index_number": "index_number",
        "court": "court",
        "case_status": "case_status",
        "notes": "notes",
    }
    for flat_key, v2_key in _case_key_map.items():
        if flat_key in row_data:
            case_fields[v2_key] = row_data[flat_key]

    case_fields["firm_name"] = firm_name

    # Map flat keys to appearance fields
    _app_key_map = {
        "appearance_date": "appearance_date",
        "outcome": "outcome",
        "charge_amount": "charge_amount",
        "invoice_number": "invoice_number",
        "invoice_sent_date": "invoice_sent_date",
        "paid_status": "paid_status",
        "payment_date": "payment_date",
        "payment_notes": "payment_notes",
        "outcome_notes": "outcome_notes",
        "docs_generated_at": "docs_generated_at",
        "draft_created_at": "draft_created_at",
    }
    for flat_key, v2_key in _app_key_map.items():
        if flat_key in row_data:
            app_fields[v2_key] = row_data[flat_key]

    return case_fields, app_fields


# ── Create ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _write_sheet_headers(ws, columns: list[str]) -> None:
    """Write styled header row to a worksheet."""
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col_idx, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(name) + 4, 14)
    ws.freeze_panes = "A2"


def create_workbook(firm_name: str, overwrite: bool = False) -> Path:
    """Create a new master dataset with v2 two-sheet format (cases + appearances)."""
    from src.file_lock import FirmFileLock

    path = dataset_path(firm_name)

    if path.exists() and not overwrite:
        raise FileExistsError(
            f"Dataset already exists: {path}\n"
            "Use --force to overwrite (this will erase all data)."
        )

    with FirmFileLock(firm_name):
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Sheet 1: cases
        ws_cases = wb.active
        ws_cases.title = "cases"
        _write_sheet_headers(ws_cases, CASE_COLUMNS)

        # Sheet 2: appearances
        ws_app = wb.create_sheet("appearances")
        _write_sheet_headers(ws_app, APPEARANCE_COLUMNS)

        wb.save(path)
    return path


def create_all_workbooks(config: dict | None = None, overwrite: bool = False) -> list[Path]:
    """Create master_cases.xlsx for every firm in config. Returns list of created paths."""
    created: list[Path] = []
    for name in all_firm_names(config):
        path = create_workbook(name, overwrite=overwrite)
        created.append(path)
    return created


# ── Load ──────────────────────────────────────────────────────────────


def _load_v1(wb) -> list[dict]:
    """Load rows from a v1 single-sheet workbook."""
    ws = wb["cases"]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def _load_v2_merged(wb) -> list[dict]:
    """Load and join cases + appearances into backward-compatible merged dicts."""
    # Build case lookup by case_id
    ws_cases = wb["cases"]
    case_headers = [cell.value for cell in ws_cases[1]]
    cases_by_id: dict[str, dict] = {}
    for row in ws_cases.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        case_dict = dict(zip(case_headers, row))
        cid = case_dict.get("case_id")
        if cid:
            cases_by_id[str(cid)] = case_dict

    # Load appearances and merge
    ws_app = wb["appearances"]
    app_headers = [cell.value for cell in ws_app[1]]
    merged_rows: list[dict] = []
    for row in ws_app.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        app_dict = dict(zip(app_headers, row))
        case_id = str(app_dict.get("case_id", ""))
        case_dict = cases_by_id.get(case_id, {})
        merged_rows.append(_merge_case_appearance(case_dict, app_dict))

    return merged_rows


def load_dataset(firm_name: str) -> list[dict]:
    """Load all rows from a firm's dataset as a list of dicts.

    Automatically detects v1 (single-sheet) vs v2 (two-sheet) format.
    Always returns backward-compatible dicts with all COLUMNS keys.
    """
    path = dataset_path(firm_name)

    if not path.exists():
        raise FileNotFoundError(
            f"Dataset not found: {path}\n"
            f"Run 'python -m src.main init-dataset' first."
        )

    wb = load_workbook(path)

    if _is_v2_format(wb):
        rows = _load_v2_merged(wb)
    else:
        rows = _load_v1(wb)

    wb.close()
    return rows


# ── Validate ──────────────────────────────────────────────────────────


def _validate_v1(wb, path) -> list[str]:
    """Validate a v1 single-sheet workbook."""
    errors: list[str] = []
    ws = wb["cases"]
    headers = [cell.value for cell in ws[1]]

    missing_cols = [c for c in COLUMNS if c not in headers]
    if missing_cols:
        errors.append(f"Missing columns: {missing_cols}")
        return errors

    seen_keys: set[tuple] = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))
        if all(v is None for v in row):
            continue
        errors.extend(_validate_row_common(row_dict, row_num))
        key = (
            str(row_dict.get("index_number", "")).strip().lower(),
            str(row_dict.get("appearance_date", "")),
        )
        if key in seen_keys:
            errors.append(f"Row {row_num}: duplicate key {key}")
        seen_keys.add(key)

    return errors


def _validate_v2(wb, path) -> list[str]:
    """Validate a v2 two-sheet workbook (cases + appearances + referential integrity)."""
    errors: list[str] = []

    # Validate cases sheet
    ws_cases = wb["cases"]
    case_headers = [cell.value for cell in ws_cases[1]]
    missing_case_cols = [c for c in CASE_COLUMNS if c not in case_headers]
    if missing_case_cols:
        errors.append(f"Cases sheet: missing columns {missing_case_cols}")
        return errors

    case_ids: set[str] = set()
    seen_index_numbers: set[str] = set()
    for row_num, row in enumerate(ws_cases.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        row_dict = dict(zip(case_headers, row))

        cid = row_dict.get("case_id")
        if not cid or str(cid).strip() == "":
            errors.append(f"Cases row {row_num}: missing case_id")
        else:
            cid_str = str(cid).strip()
            if cid_str in case_ids:
                errors.append(f"Cases row {row_num}: duplicate case_id '{cid_str}'")
            case_ids.add(cid_str)

        idx = row_dict.get("index_number")
        if not idx or str(idx).strip() == "":
            errors.append(f"Cases row {row_num}: missing index_number")
        else:
            idx_key = str(idx).strip().lower()
            if idx_key in seen_index_numbers:
                errors.append(f"Cases row {row_num}: duplicate index_number '{idx}'")
            seen_index_numbers.add(idx_key)

        cs = row_dict.get("case_status")
        if cs is not None and str(cs).strip() != "":
            if str(cs).strip() not in VALID_CASE_STATUSES:
                errors.append(
                    f"Cases row {row_num}: case_status '{cs}' not in {VALID_CASE_STATUSES}"
                )

    # Validate appearances sheet
    ws_app = wb["appearances"]
    app_headers = [cell.value for cell in ws_app[1]]
    missing_app_cols = [c for c in APPEARANCE_COLUMNS if c not in app_headers]
    if missing_app_cols:
        errors.append(f"Appearances sheet: missing columns {missing_app_cols}")
        return errors

    seen_app_ids: set[str] = set()
    seen_keys: set[tuple] = set()
    for row_num, row in enumerate(ws_app.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        row_dict = dict(zip(app_headers, row))

        # appearance_id uniqueness
        aid = row_dict.get("appearance_id")
        if not aid or str(aid).strip() == "":
            errors.append(f"Appearances row {row_num}: missing appearance_id")
        else:
            aid_str = str(aid).strip()
            if aid_str in seen_app_ids:
                errors.append(f"Appearances row {row_num}: duplicate appearance_id '{aid_str}'")
            seen_app_ids.add(aid_str)

        # FK check: case_id must exist in cases sheet
        fk = row_dict.get("case_id")
        if not fk or str(fk).strip() == "":
            errors.append(f"Appearances row {row_num}: missing case_id")
        elif str(fk).strip() not in case_ids:
            errors.append(
                f"Appearances row {row_num}: case_id '{fk}' not found in cases sheet"
            )

        # Validate common fields via merged dict approach
        # Build a pseudo-merged dict for validation
        ad = row_dict.get("appearance_date")
        if ad is not None:
            if isinstance(ad, datetime):
                pass
            elif isinstance(ad, date):
                pass
            elif isinstance(ad, str):
                try:
                    datetime.strptime(ad, "%Y-%m-%d")
                except ValueError:
                    errors.append(
                        f"Appearances row {row_num}: appearance_date '{ad}' is not YYYY-MM-DD"
                    )

        amt = row_dict.get("charge_amount")
        if amt is not None and not isinstance(amt, (int, float)):
            try:
                float(amt)
            except (ValueError, TypeError):
                errors.append(
                    f"Appearances row {row_num}: charge_amount '{amt}' is not a number"
                )

        ps = row_dict.get("paid_status")
        if ps is not None and str(ps).strip() != "":
            if str(ps).strip() not in VALID_PAID_STATUSES:
                errors.append(
                    f"Appearances row {row_num}: paid_status '{ps}' not in {VALID_PAID_STATUSES}"
                )

        # Unique key check: (case_id, appearance_date)
        key = (
            str(fk or "").strip(),
            str(ad or ""),
        )
        if key in seen_keys:
            errors.append(f"Appearances row {row_num}: duplicate key (case_id, appearance_date) {key}")
        seen_keys.add(key)

    return errors


def _validate_row_common(row_dict: dict, row_num: int) -> list[str]:
    """Validate fields common to v1 flat rows. Returns list of errors."""
    errors: list[str] = []

    for col in REQUIRED_COLUMNS:
        if row_dict.get(col) is None or str(row_dict[col]).strip() == "":
            errors.append(f"Row {row_num}: missing required field '{col}'")

    ad = row_dict.get("appearance_date")
    if ad is not None:
        if isinstance(ad, datetime):
            pass
        elif isinstance(ad, date):
            pass
        elif isinstance(ad, str):
            try:
                datetime.strptime(ad, "%Y-%m-%d")
            except ValueError:
                errors.append(
                    f"Row {row_num}: appearance_date '{ad}' is not YYYY-MM-DD"
                )

    amt = row_dict.get("charge_amount")
    if amt is not None and not isinstance(amt, (int, float)):
        try:
            float(amt)
        except (ValueError, TypeError):
            errors.append(
                f"Row {row_num}: charge_amount '{amt}' is not a number"
            )

    cs = row_dict.get("case_status")
    if cs is not None and str(cs).strip() != "":
        if str(cs).strip() not in VALID_CASE_STATUSES:
            errors.append(
                f"Row {row_num}: case_status '{cs}' not in {VALID_CASE_STATUSES}"
            )

    ps = row_dict.get("paid_status")
    if ps is not None and str(ps).strip() != "":
        if str(ps).strip() not in VALID_PAID_STATUSES:
            errors.append(
                f"Row {row_num}: paid_status '{ps}' not in {VALID_PAID_STATUSES}"
            )

    return errors


def validate_dataset(firm_name: str) -> list[str]:
    """Validate a firm's dataset file. Returns list of error messages (empty = OK).

    Automatically detects v1 vs v2 format.
    """
    path = dataset_path(firm_name)
    errors: list[str] = []

    if not path.exists():
        return [f"Dataset file not found: {path}"]

    wb = load_workbook(path)

    if _is_v2_format(wb):
        errors = _validate_v2(wb, path)
    elif "cases" in wb.sheetnames:
        errors = _validate_v1(wb, path)
    else:
        errors = ["Missing required sheet 'cases'"]

    wb.close()
    return errors


# ── Lookup ────────────────────────────────────────────────────────────


def find_row_by_key(
    firm_name: str,
    index_number: str,
    appearance_date: str | date,
    rows: list[dict] | None = None,
) -> dict | None:
    """Find a row by its unique key within a firm's dataset. Returns the dict or None."""
    if rows is None:
        rows = load_dataset(firm_name)

    target_idx = index_number.strip().lower()
    target_date = str(appearance_date)

    for row in rows:
        if (
            str(row.get("index_number", "")).strip().lower() == target_idx
            and str(row.get("appearance_date", "")) == target_date
        ):
            return row

    return None


# ── Upsert ────────────────────────────────────────────────────────────


def _match_key(ws, headers: list[str], index_number: str, appearance_date: str) -> int | None:
    """Return the Excel row number of a matching key, or None."""
    idx_col = headers.index("index_number")
    date_col = headers.index("appearance_date")
    target_idx = index_number.strip().lower()
    target_date = str(appearance_date)

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_idx = str(row[idx_col] or "").strip().lower()
        row_date = str(row[date_col] or "")
        if row_idx == target_idx and row_date == target_date:
            return row_num
    return None


def _match_appearance(ws_app, app_headers: list[str], case_id: str, appearance_date: str) -> int | None:
    """Return the Excel row number of a matching appearance, or None."""
    cid_col = app_headers.index("case_id")
    date_col = app_headers.index("appearance_date")
    target_date = str(appearance_date)

    for row_num, row in enumerate(ws_app.iter_rows(min_row=2, values_only=True), start=2):
        row_cid = str(row[cid_col] or "").strip()
        row_date = str(row[date_col] or "")
        if row_cid == case_id and row_date == target_date:
            return row_num
    return None


def _match_case_by_index(ws_cases, case_headers: list[str], index_number: str) -> int | None:
    """Return the Excel row number of a case matching index_number, or None."""
    idx_col = case_headers.index("index_number")
    target = index_number.strip().lower()

    for row_num, row in enumerate(ws_cases.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[idx_col] or "").strip().lower() == target:
            return row_num
    return None


def _upsert_v1(wb, path, row_data: dict) -> str:
    """Upsert into a v1 single-sheet workbook."""
    ws = wb["cases"]
    headers = [cell.value for cell in ws[1]]

    idx_num = str(row_data.get("index_number", ""))
    app_date = str(row_data.get("appearance_date", ""))
    existing_row = _match_key(ws, headers, idx_num, app_date)

    if existing_row is not None:
        for col_name, value in row_data.items():
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                ws.cell(row=existing_row, column=col_idx, value=value)
        wb.save(path)
        wb.close()
        return "updated"
    else:
        new_row = [row_data.get(col) for col in headers]
        ws.append(new_row)
        wb.save(path)
        wb.close()
        return "inserted"


def _upsert_v2(wb, path, firm_name: str, row_data: dict) -> str:
    """Upsert into a v2 two-sheet workbook. Splits data across cases + appearances."""
    case_fields, app_fields = _split_row_data(row_data, firm_name)

    ws_cases = wb["cases"]
    case_headers = [cell.value for cell in ws_cases[1]]
    ws_app = wb["appearances"]
    app_headers = [cell.value for cell in ws_app[1]]

    idx_num = str(row_data.get("index_number", ""))
    app_date = str(row_data.get("appearance_date", ""))

    # Find or create case
    case_row_num = _match_case_by_index(ws_cases, case_headers, idx_num)
    if case_row_num is not None:
        # Update existing case fields
        case_id_col = case_headers.index("case_id") + 1
        case_id = str(ws_cases.cell(row=case_row_num, column=case_id_col).value)
        for col_name, value in case_fields.items():
            if col_name in case_headers:
                col_idx = case_headers.index(col_name) + 1
                ws_cases.cell(row=case_row_num, column=col_idx, value=value)
    else:
        # Insert new case
        case_id = str(uuid.uuid4())
        case_fields["case_id"] = case_id
        if "date_added" not in case_fields:
            case_fields["date_added"] = date.today().isoformat()
        new_case_row = [case_fields.get(col) for col in case_headers]
        ws_cases.append(new_case_row)

    # Find or create appearance
    app_row_num = _match_appearance(ws_app, app_headers, case_id, app_date)
    if app_row_num is not None:
        # Update existing appearance
        for col_name, value in app_fields.items():
            if col_name in app_headers:
                col_idx = app_headers.index(col_name) + 1
                ws_app.cell(row=app_row_num, column=col_idx, value=value)
        wb.save(path)
        wb.close()
        return "updated"
    else:
        # Insert new appearance
        app_fields["appearance_id"] = str(uuid.uuid4())
        app_fields["case_id"] = case_id
        new_app_row = [app_fields.get(col) for col in app_headers]
        ws_app.append(new_app_row)
        wb.save(path)
        wb.close()
        return "inserted"


def upsert_row(firm_name: str, row_data: dict, _hold_lock: bool = True) -> str:
    """Insert or update a row in a firm's dataset.

    row_data keys should match COLUMNS (extras are ignored, missing become None).
    Uses (index_number, appearance_date) as the unique key.

    Automatically handles v1 and v2 formats:
    - v1: writes to single 'cases' sheet (legacy)
    - v2: splits data across 'cases' + 'appearances' sheets

    _hold_lock: if True (default), acquires per-firm lock. Set to False when
    the caller already holds the lock (e.g. batch imports).

    Returns "inserted" or "updated".
    """
    from src.file_lock import FirmFileLock

    path = dataset_path(firm_name)

    if not path.exists():
        raise FileNotFoundError(
            f"Dataset not found: {path}\n"
            "Run 'python -m src.main init-dataset' first."
        )

    lock = FirmFileLock(firm_name) if _hold_lock else contextlib.nullcontext()

    with lock:
        wb = load_workbook(path)

        if _is_v2_format(wb):
            return _upsert_v2(wb, path, firm_name, row_data)
        else:
            return _upsert_v1(wb, path, row_data)


# ── Query ─────────────────────────────────────────────────────────────


def _to_date(val) -> date | None:
    """Coerce a value to a date object, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).split(" ")[0], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def query_by_date_range(
    firm_name: str,
    start: date,
    end: date,
    rows: list[dict] | None = None,
) -> list[dict]:
    """Return rows whose appearance_date falls within [start, end] inclusive."""
    if rows is None:
        rows = load_dataset(firm_name)

    result = []
    for row in rows:
        d = _to_date(row.get("appearance_date"))
        if d is not None and start <= d <= end:
            result.append(row)

    # Sort by appearance_date
    result.sort(key=lambda r: _to_date(r.get("appearance_date")) or date.min)
    return result


def week_range(ref_date: date) -> tuple[date, date]:
    """Return (monday, friday) of the business week containing ref_date."""
    monday = ref_date - timedelta(days=ref_date.weekday())  # weekday: Mon=0
    friday = monday + timedelta(days=4)
    return monday, friday


def month_range(year: int, month: int) -> tuple[date, date]:
    """Return (first_day, last_day) of the given month."""
    first = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    last = date(year, month, last_day)
    return first, last
