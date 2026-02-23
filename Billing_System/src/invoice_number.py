"""Invoice numbering generator (per-firm, persistent counter).

Each firm has a counter file at  data/{FirmName}/invoice_counter.json
Format from config:  {initials}{year}{number:03d}   e.g. AL2026001
Counter resets each year when yearly_reset is true.
"""

import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src.config import CONFIG_PATH, load_config, get_firm
from src.dataset import dataset_path, get_data_root


def _counter_path(firm_name: str) -> Path:
    return get_data_root() / "invoice" / firm_name / "invoice_counter.json"


def _load_counter(firm_name: str) -> dict:
    """Load or initialise the counter for a firm."""
    path = _counter_path(firm_name)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"year": date.today().year, "last_number": 0}


def _save_counter(firm_name: str, counter: dict) -> None:
    path = _counter_path(firm_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(counter, f, indent=2)


def next_invoice_number(firm_name: str, config: dict | None = None) -> str:
    """Generate the next invoice number for a firm and persist the counter.

    Uses the format and yearly_reset settings from config.
    """
    if config is None:
        config = load_config()

    firm = get_firm(firm_name, config)
    initials = firm["initials"]
    numbering = config.get("invoice_numbering", {})
    fmt = numbering.get("format", "{initials}{year}{number:03d}")
    yearly_reset = numbering.get("yearly_reset", True)

    counter = _load_counter(firm_name)
    current_year = date.today().year

    if yearly_reset and counter["year"] != current_year:
        counter["year"] = current_year
        counter["last_number"] = 0

    counter["last_number"] += 1
    _save_counter(firm_name, counter)

    return fmt.format(
        initials=initials,
        year=counter["year"],
        number=counter["last_number"],
    )


def assign_invoice_numbers(firm_name: str, config: dict | None = None) -> list[str]:
    """Assign invoice numbers to all rows in a firm's dataset that lack one.

    Returns list of newly assigned invoice numbers.
    """
    from src.file_lock import FirmFileLock

    if config is None:
        config = load_config()

    path = dataset_path(firm_name)
    if not path.exists():
        raise FileNotFoundError(
            f"Dataset not found: {path}\n"
            "Run 'python -m src.main init-dataset' first."
        )

    with FirmFileLock(firm_name):
        wb = load_workbook(path)
        ws = wb["cases"]
        headers = [cell.value for cell in ws[1]]
        inv_col = headers.index("invoice_number") + 1  # 1-based

        assigned: list[str] = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            current_inv = row[inv_col - 1]  # 0-based for tuple
            if current_inv is None or str(current_inv).strip() == "":
                inv_num = next_invoice_number(firm_name, config)
                ws.cell(row=row_num, column=inv_col, value=inv_num)
                assigned.append(inv_num)

        wb.save(path)
        wb.close()
    return assigned
