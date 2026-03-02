"""Payment update workflow — mark invoices as Paid, Unpaid, or Partial.

Updates paid_status and payment_date in the firm's dataset.
Logs every change to invoice/{FirmName}/payment_log.csv for audit.
"""

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from src.dataset import dataset_path, get_data_root, load_dataset, _is_v2_format


# ── Lookup by invoice number ─────────────────────────────────────────

def find_by_invoice_number(
    firm_name: str,
    invoice_number: str,
) -> dict | None:
    """Find a case row by its invoice_number. Returns the dict or None."""
    rows = load_dataset(firm_name)
    target = invoice_number.strip()
    for row in rows:
        if str(row.get("invoice_number") or "").strip() == target:
            return row
    return None


def _match_invoice_row(ws, headers: list[str], invoice_number: str) -> int | None:
    """Return the Excel row number matching the invoice_number, or None."""
    inv_col = headers.index("invoice_number")
    target = invoice_number.strip()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        val = str(row[inv_col] or "").strip()
        if val == target:
            return row_num
    return None


# ── Update payment ───────────────────────────────────────────────────

VALID_STATUSES = {"Paid", "Unpaid", "Partial"}


def mark_payment(
    firm_name: str,
    invoice_number: str,
    status: str,
    payment_date: date | str | None = None,
    notes: str | None = None,
) -> dict:
    """Update paid_status (and optionally payment_date, notes) for a case.

    Handles both v1 ('cases' sheet) and v2 ('appearances' sheet) formats.
    Returns the updated case dict.
    Raises ValueError if invoice not found or invalid status.
    """
    from src.file_lock import FirmFileLock

    if status not in VALID_STATUSES:
        raise ValueError(f"Invalid status '{status}'. Must be one of: {VALID_STATUSES}")

    path = dataset_path(firm_name)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found: {path}")

    with FirmFileLock(firm_name):
        wb = load_workbook(path)

        # Pick the right sheet and notes column name
        if _is_v2_format(wb):
            ws = wb["appearances"]
            notes_col_name = "payment_notes"
        else:
            ws = wb["cases"]
            notes_col_name = "notes"

        headers = [cell.value for cell in ws[1]]

        row_num = _match_invoice_row(ws, headers, invoice_number)
        if row_num is None:
            wb.close()
            raise ValueError(
                f"Invoice '{invoice_number}' not found in {firm_name}'s dataset."
            )

        # Read current row for audit log
        row_values = [ws.cell(row=row_num, column=c + 1).value for c in range(len(headers))]
        old_row = dict(zip(headers, row_values))

        # Update paid_status
        ps_col = headers.index("paid_status") + 1
        ws.cell(row=row_num, column=ps_col, value=status)

        # Update payment_date
        pd_col = headers.index("payment_date") + 1
        if payment_date is not None:
            if isinstance(payment_date, str):
                payment_date = datetime.strptime(payment_date, "%Y-%m-%d").date()
            ws.cell(row=row_num, column=pd_col, value=payment_date)
        elif status == "Paid" and old_row.get("payment_date") is None:
            # Auto-set payment_date to today if marking Paid and no date provided
            ws.cell(row=row_num, column=pd_col, value=date.today())

        # Update notes if provided
        if notes is not None:
            nc = headers.index(notes_col_name) + 1
            ws.cell(row=row_num, column=nc, value=notes)

        wb.save(path)
        wb.close()

        # Write audit log
        # For audit, map back to case_caption from the merged view
        _write_audit_log(firm_name, invoice_number, old_row, status, payment_date)

    # Re-read to return clean merged dict
    return find_by_invoice_number(firm_name, invoice_number) or {}


# ── Audit log ────────────────────────────────────────────────────────

def _audit_log_path(firm_name: str) -> Path:
    return get_data_root() / "invoice" / firm_name / "payment_log.csv"


def _write_audit_log(
    firm_name: str,
    invoice_number: str,
    old_row: dict,
    new_status: str,
    payment_date: date | str | None,
) -> None:
    """Append a line to the firm's payment audit log."""
    log_path = _audit_log_path(firm_name)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    file_exists = log_path.exists()

    # old_row may come from v2 appearances sheet (no case_caption) or v1 (has it)
    caption = old_row.get("case_caption") or old_row.get("caption", "")

    with open(log_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow([
                "timestamp", "invoice_number", "case_caption",
                "old_status", "new_status", "payment_date",
            ])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            invoice_number,
            caption,
            old_row.get("paid_status", ""),
            new_status,
            str(payment_date) if payment_date else "",
        ])
