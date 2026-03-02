"""Migrate firm datasets from v1 (flat single-sheet) to v2 (cases + appearances).

Safety:
  - Holds FirmFileLock during migration
  - Renames old sheet to 'cases_v1_backup' (preserved in the workbook)
  - Supports --dry-run mode (reports what would change without writing)
  - Validates result after migration

Usage:
  python -m src.main migrate-v2 [--firm FIRM] [--dry-run]
"""

import uuid
from collections import defaultdict
from datetime import date, datetime

from openpyxl import load_workbook

from src.dataset import (
    APPEARANCE_COLUMNS,
    CASE_COLUMNS,
    _is_v2_format,
    _write_sheet_headers,
    dataset_path,
    validate_dataset,
    all_firm_names,
)


def _to_str(val) -> str:
    """Coerce a value to a string for grouping."""
    if val is None:
        return ""
    return str(val).strip()


def migrate_firm(firm_name: str, dry_run: bool = False) -> dict:
    """Migrate a single firm's dataset from v1 to v2.

    Returns a summary dict with keys:
      - firm: str
      - status: "migrated" | "already_v2" | "skipped" | "error"
      - cases_created: int
      - appearances_created: int
      - errors: list[str]
      - message: str
    """
    from src.file_lock import FirmFileLock

    path = dataset_path(firm_name)
    result = {
        "firm": firm_name,
        "status": "skipped",
        "cases_created": 0,
        "appearances_created": 0,
        "errors": [],
        "message": "",
    }

    if not path.exists():
        result["status"] = "skipped"
        result["message"] = f"Dataset not found: {path}"
        return result

    with FirmFileLock(firm_name):
        wb = load_workbook(path)

        # Already v2?
        if _is_v2_format(wb):
            wb.close()
            result["status"] = "already_v2"
            result["message"] = f"{firm_name}: already in v2 format."
            return result

        # Must have a 'cases' sheet (v1)
        if "cases" not in wb.sheetnames:
            wb.close()
            result["status"] = "error"
            result["errors"] = [f"No 'cases' sheet found in {path}"]
            result["message"] = f"{firm_name}: missing 'cases' sheet."
            return result

        ws_old = wb["cases"]
        old_headers = [cell.value for cell in ws_old[1]]

        # Read all v1 rows
        v1_rows: list[dict] = []
        for row in ws_old.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            v1_rows.append(dict(zip(old_headers, row)))

        if dry_run:
            # Count what would be created
            grouped = defaultdict(list)
            for row in v1_rows:
                idx = _to_str(row.get("index_number"))
                grouped[idx].append(row)

            result["status"] = "dry_run"
            result["cases_created"] = len(grouped)
            result["appearances_created"] = len(v1_rows)
            result["message"] = (
                f"{firm_name}: DRY RUN — would create {len(grouped)} case(s) "
                f"and {len(v1_rows)} appearance(s) from {len(v1_rows)} v1 row(s)."
            )
            wb.close()
            return result

        # Group v1 rows by index_number to create case records
        grouped = defaultdict(list)
        for row in v1_rows:
            idx = _to_str(row.get("index_number"))
            grouped[idx].append(row)

        # Build case rows and appearance rows
        case_rows: list[dict] = []
        appearance_rows: list[dict] = []

        for index_number, rows in grouped.items():
            case_id = str(uuid.uuid4())

            # Take case-level fields from the most recent appearance
            rows_sorted = sorted(
                rows,
                key=lambda r: str(r.get("appearance_date", "")),
                reverse=True,
            )
            most_recent = rows_sorted[0]

            case_row = {
                "case_id": case_id,
                "firm_name": firm_name,
                "caption": most_recent.get("case_caption"),
                "index_number": index_number,
                "court": most_recent.get("court"),
                "case_status": most_recent.get("case_status"),
                "date_added": date.today().isoformat(),
                "notes": most_recent.get("notes"),
            }
            case_rows.append(case_row)

            # Create appearance rows for each v1 row in this group
            for row in rows:
                app_row = {
                    "appearance_id": str(uuid.uuid4()),
                    "case_id": case_id,
                    "appearance_date": row.get("appearance_date"),
                    "outcome": row.get("outcome"),
                    "outcome_notes": None,
                    "charge_amount": row.get("charge_amount"),
                    "invoice_number": row.get("invoice_number"),
                    "invoice_sent_date": row.get("invoice_sent_date"),
                    "docs_generated_at": None,
                    "draft_created_at": None,
                    "paid_status": row.get("paid_status"),
                    "payment_date": row.get("payment_date"),
                    "payment_notes": None,
                }
                appearance_rows.append(app_row)

        # Rename old sheet to backup
        ws_old.title = "cases_v1_backup"

        # Create new cases sheet
        ws_cases = wb.create_sheet("cases", 0)  # insert at position 0
        _write_sheet_headers(ws_cases, CASE_COLUMNS)
        for case in case_rows:
            ws_cases.append([case.get(col) for col in CASE_COLUMNS])

        # Create new appearances sheet
        ws_app = wb.create_sheet("appearances", 1)  # insert at position 1
        _write_sheet_headers(ws_app, APPEARANCE_COLUMNS)
        for app in appearance_rows:
            ws_app.append([app.get(col) for col in APPEARANCE_COLUMNS])

        wb.save(path)
        wb.close()

    # Validate the result
    errors = validate_dataset(firm_name)

    result["status"] = "migrated"
    result["cases_created"] = len(case_rows)
    result["appearances_created"] = len(appearance_rows)
    result["errors"] = errors
    result["message"] = (
        f"{firm_name}: migrated {len(v1_rows)} v1 row(s) -> "
        f"{len(case_rows)} case(s) + {len(appearance_rows)} appearance(s)."
    )
    if errors:
        result["message"] += f"\n  Validation warnings: {len(errors)}"
        for err in errors:
            result["message"] += f"\n    - {err}"

    return result


def migrate_all_firms(config: dict | None = None, dry_run: bool = False) -> dict:
    """Migrate all firms from v1 to v2.

    Returns a summary dict with keys:
      - total: int
      - migrated: int
      - already_v2: int
      - skipped: int
      - errors: int
      - firms: list[dict]  (per-firm results)
    """
    firms = all_firm_names(config)
    summary = {
        "total": len(firms),
        "migrated": 0,
        "already_v2": 0,
        "skipped": 0,
        "errors": 0,
        "firms": [],
    }

    for name in firms:
        result = migrate_firm(name, dry_run=dry_run)
        summary["firms"].append(result)

        if result["status"] == "migrated" or result["status"] == "dry_run":
            summary["migrated"] += 1
        elif result["status"] == "already_v2":
            summary["already_v2"] += 1
        elif result["status"] == "error":
            summary["errors"] += 1
        else:
            summary["skipped"] += 1

    return summary
